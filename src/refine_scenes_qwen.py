import os
import sys
import json
import datetime
import glob
import openpyxl
from openai import OpenAI

# Add project root to sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
sys.path.append(project_root)

from src.utils import load_config

def get_product_info(file_path):
    """
    Reads product name and detail from ec_product.xlsx.
    """
    if not os.path.exists(file_path):
        return None, None
        
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        headers = {}
        for cell in sheet[1]:
            if cell.value:
                headers[cell.column] = cell.value
                
        name_col = None
        detail_col = None
        
        for col_idx, header_val in headers.items():
            if header_val == 'name':
                name_col = col_idx
            elif header_val == 'detail':
                detail_col = col_idx
                
        if sheet.max_row >= 2:
            name = sheet.cell(row=2, column=name_col).value if name_col else None
            detail = sheet.cell(row=2, column=detail_col).value if detail_col else None
            return name, detail
            
    except Exception as e:
        print(f"Error reading product info: {e}")
        
    return None, None

def get_latest_scene_summary(outputs_dir):
    # Match both original and refined to find the absolute latest source if needed,
    # but user said "refine THE generated json", implying the one from previous step.
    # Previous step generated 'scene_summary_*.json'.
    files = glob.glob(os.path.join(outputs_dir, "scene_summary_*.json"))
    if not files:
        return None
    return max(files, key=os.path.getctime)

def refine_scenes():
    config = load_config()
    api_key = config.get("api_keys", {}).get("dashscope") or os.environ.get("DASHSCOPE_API_KEY")
    if not api_key:
        print("Error: DASHSCOPE_API_KEY not found.")
        return

    outputs_dir = os.path.join(project_root, "outputs")
    latest_file = get_latest_scene_summary(outputs_dir)
    
    if not latest_file:
        print("No scene_summary JSON file found in outputs directory.")
        return
        
    print(f"Reading from: {latest_file}")
    with open(latest_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Load product info
    excel_path = os.path.join(project_root, "ec_product.xlsx")
    product_name, product_desc = get_product_info(excel_path)
    print(f"Product: {product_name}")

    # Prepare example scene from input data for prompt
    first_scene_str = ""
    if "scenes" in data and isinstance(data["scenes"], list) and len(data["scenes"]) > 0:
        example_scene = data["scenes"][0].copy()
        # Ensure example has 'source' key for the prompt example
        if "source" not in example_scene:
            example_scene["source"] = "optimized_original"
        first_scene_str = json.dumps(example_scene, ensure_ascii=False, indent=8)
    else:
        # Fallback
        first_scene_str = json.dumps({
            "id": 1, 
            "scene_name": "示例", 
            "source": "optimized_original"
        }, ensure_ascii=False, indent=8)

    # Construct Prompt
    prompt = f"""
        ### 角色设定
        你是一位资深电商运营与海报摄影导演，将帮助我生成商品展示海报的“场景短语”。所有描述必须用于**静态**的商品展示海报，画面必须是静止的瞬间，而非动态视频。

        ### 商品信息
        - 名称：{product_name}
        - 品牌风格：如有特定风格请在此处体现（如“简约”、“复古”、“科技感”等）
        - 目标人群：如有特定人群请在此处体现（如“年轻女性”、“商务人士”等）

        ### 输入 JSON 数据
        以下是需要优化和泛化的原始 JSON 数据，请基于此进行操作：
        {json.dumps(data, ensure_ascii=False, indent=2)}

        ### 任务 1：优化现有场景
        1. 遍历输入 JSON 中 `scenes` 列表，对每个场景的 `description`（场景描述）、`surrounding_objects`（周围物体）、`details`（细节展示）和 `selling_point`（卖点展示）进行优化。
        2. **更清晰**：用具象名词和视觉动作替代抽象形容词，使描述直观且易于想象。
        3. **更简单**：简化句子结构，使其符合 Stable Diffusion 或 Midjourney 风格的提示词，但保持中文表达。
        4. **画面化**：确保 `details` 和 `selling_point` 的内容描述的是画面元素或视觉特写，而非功能说明。

        ### 任务 2：扩展新场景
        1. 结合产品名称、描述及输入 JSON 中已有的场景内容，发散思维但保持合理，在此基础上额外生成 5 个新的静态商品展示场景。
        2. 泛化维度包括：使用场景（室内、户外、办公、休闲等）、时间/季节（春夏秋冬、不同时段光线）、背景主题（现代都市、自然风光、简约摄影棚等）、情绪氛围（温馨、活力、专业、奢华等）。
        3. 避免与原有场景重复，也不要天马行空；每个新场景应突出商品特色并符合品牌风格。
        4. 为新场景生成唯一的 `id`（在现有最大 `id` 基础上顺延）和简洁独特的 `scene_name`。

        ### 输出格式
        请将优化后的原有场景和新增的 5 个场景合并为一个新的 `scenes` 列表，每个场景对象必须包含以下字段：
        - `id`: 数字标识，不重复。
        - `scene_name`: 概括场景主题的名称。
        - `description`: 对场景画面的整体描述，静态的画面，不要有产品本身或者使用产品的描述存在。
        - `surrounding_objects`: 出现在画面中的其他物体，用逗号分隔。
        - `details`: 需要突出的视觉细节，如光影、材质特写等，禁止出现跟产品相关或者产品替代物的描述。
        - `selling_point`: 突出卖点的画面展示方式。
        - `source`: 对原有场景填入 "optimized"，对新增场景填入 "new"。

        ### 示例场景（仅供参考，用于理解结构）
        {first_scene_str}

        ### 输出要求
        - 请直接返回修改后的完整 JSON 数据，不要包含任何 Markdown 标记（如 ```json）。
        - 输出的根对象应该至少包含 `scenes` 字段，且结构与示例一致。
    """

    #print(f"Prompt:\n{prompt}\n\n")
    
    client = OpenAI(
        api_key=api_key,
        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
    )

    print("Calling Qwen (qwen-plus) for refinement...")
    try:
        completion = client.chat.completions.create(
            model="qwen-plus",
            messages=[
                {"role": "system", "content": "You are a helpful assistant that outputs valid JSON only."},
                {"role": "user", "content": prompt}
            ],
        )
        
        content = completion.choices[0].message.content.strip()
        
        # Simple cleanup if markdown is present
        if content.startswith("```json"):
            content = content[7:]
        elif content.startswith("```"):
            content = content[3:]
        if content.endswith("```"):
            content = content[:-3]
            
        refined_data = json.loads(content)
        
        # Save to file
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"refined_scene_summary_{timestamp}.json"
        output_path = os.path.join(outputs_dir, output_filename)
        
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(refined_data, f, indent=4, ensure_ascii=False)
            
        print(f"\nRefined JSON saved to: {output_path}")
        print(json.dumps(refined_data, indent=4, ensure_ascii=False))

    except Exception as e:
        print(f"Error calling Qwen: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    refine_scenes()
