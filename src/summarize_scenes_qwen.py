import os
import sys
import base64
import io
import openpyxl
import csv
import json
import datetime
from PIL import Image
from openai import OpenAI

# Add project root to sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
sys.path.append(project_root)

from src.utils import load_config

def encode_image(image_path):
    """
    Resizes image to 512x512 and encodes to base64.
    """
    try:
        with Image.open(image_path) as img:
            img = img.convert("RGB")
            img = img.resize((512, 512))
            buffered = io.BytesIO()
            img.save(buffered, format="JPEG")
            return base64.b64encode(buffered.getvalue()).decode('utf-8')
    except Exception as e:
        print(f"Error processing image {image_path}: {e}")
        return None

def get_product_info_with_samples(file_path):
    """
    Reads product name, detail, and sample_dir from ec_product.xlsx.
    """
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return None, None, None

    name = None
    detail = None
    sample_dir = None

    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        headers = {}
        for cell in sheet[1]:
            if cell.value:
                headers[cell.column] = str(cell.value).strip()
        
        name_col = None
        detail_col = None
        sample_dir_col = None
        
        for col_idx, header_val in headers.items():
            if header_val == 'name':
                name_col = col_idx
            elif header_val == 'detail':
                detail_col = col_idx
            elif header_val == 'sample_dir':
                sample_dir_col = col_idx
        
        if not name_col: print("Column 'name' not found")
        if not detail_col: print("Column 'detail' not found")
        if not sample_dir_col: print("Column 'sample_dir' not found")

        if sheet.max_row >= 2:
            if name_col:
                name = sheet.cell(row=2, column=name_col).value
            if detail_col:
                detail = sheet.cell(row=2, column=detail_col).value
            if sample_dir_col:
                sample_dir = sheet.cell(row=2, column=sample_dir_col).value
            return name, detail, sample_dir
        else:
            print("No data rows in xlsx")
                
    except Exception as e:
        print(f"Error reading file: {e}")
        return None, None, None
    
    return None, None, None

def summarize_scenes():
    config = load_config()
    api_key = config.get("api_keys", {}).get("dashscope") or os.environ.get("DASHSCOPE_API_KEY")
    if not api_key:
        print("Error: DASHSCOPE_API_KEY not found.")
        return

    excel_path = os.path.join(project_root, "ec_product.xlsx")
    print(f"Reading product info from {excel_path}...")
    product_name, product_desc, sample_dir = get_product_info_with_samples(excel_path)
    
    print(f"Product Name: {product_name}")
    print(f"Product Description: {product_desc}")
    print(f"Sample Directory: {sample_dir}")

    if not sample_dir:
        print("Sample dir not found or empty in Excel.")
        return
        
    # Resolve sample_dir path
    if not os.path.isabs(sample_dir):
        # Check if it exists relative to project root
        sample_dir_abs = os.path.join(project_root, sample_dir)
        if not os.path.exists(sample_dir_abs):
             # Check if it exists relative to current working dir (just in case)
             if os.path.exists(sample_dir):
                 sample_dir_abs = os.path.abspath(sample_dir)
        sample_dir = sample_dir_abs
        
    if not os.path.exists(sample_dir):
        print(f"Sample directory does not exist: {sample_dir}")
        return

    # Collect images
    image_contents = []
    valid_extensions = {".jpg", ".jpeg", ".png", ".webp"}
    print(f"Processing images from {sample_dir}...")
    
    files = sorted(os.listdir(sample_dir))
    for filename in files:
        ext = os.path.splitext(filename)[1].lower()
        if ext in valid_extensions:
            img_path = os.path.join(sample_dir, filename)
            base64_img = encode_image(img_path)
            if base64_img:
                image_contents.append({
                    "type": "image_url",
                    "image_url": {"url": f"data:image/jpeg;base64,{base64_img}"}
                })
                print(f"  Loaded and resized {filename}")

    if not image_contents:
        print("No valid images found in sample_dir.")
        return

    # Construct Prompt
    prompt_text = f"""
    产品名称：{product_name}
    产品描述：{product_desc}
    
    任务：
    请作为一名专业的电商视觉策划，完成以下分析并输出标准的 JSON 格式数据。

    步骤 1：判断提供的参考图片内容是否与“产品名称”和“产品描述”一致。
    步骤 2：总结或推断适合该产品的使用场景（如果图片不符，请根据产品文本描述推断场景）。

    输出格式要求（必须是合法的 JSON）：
    {{
        "is_match": boolean,  // 图片是否符合产品描述，符合为 true，不符合为 false
        "mismatch_reason": string, // 如果不符合，请说明原因；如果符合，可为空字符串
        "scene_count": integer, // 总结的场景数量
        "scenes": [ // 场景列表
            {{
                "id": integer, // 场景序号
                "scene_name": string, // 场景名称
                "description": string, // 场景描述（光线、氛围）
                "surrounding_objects": string, // 周围物体
                "details": string, // 细节展示
                "selling_point": string // 卖点描述（关联功能）
            }}
        ]
    }}

    注意：
    1. 直接返回 JSON 字符串，不要包含 ```json 或其他 Markdown 标记。
    2. 场景描述要具体，具有画面感。
    3. 即使图片不匹配，也必须根据产品文本生成场景推荐。
    """
    
    messages = [
        {
            "role": "user",
            "content": [
                {"type": "text", "text": prompt_text},
                *image_contents
            ]
        }
    ]

    client = OpenAI(
        api_key=api_key,
        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
    )

    print("\nCalling Qwen VL (qwen-vl-plus)...")
    try:
        completion = client.chat.completions.create(
            model="qwen-vl-plus",
            messages=messages,
        )
        print("\n=== Qwen Scene Summary (JSON) ===\n")
        content = completion.choices[0].message.content
        
        # Simple cleanup if markdown is present despite instructions
        if content.startswith("```json"):
            content = content[7:]
        if content.endswith("```"):
            content = content[:-3]
        
        try:
            json_data = json.loads(content)
            print(json.dumps(json_data, indent=4, ensure_ascii=False))

            # Save to file
            outputs_dir = os.path.join(project_root, "outputs")
            if not os.path.exists(outputs_dir):
                os.makedirs(outputs_dir)
            
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"scene_summary_{timestamp}.json"
            output_path = os.path.join(outputs_dir, output_filename)
            
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(json_data, f, indent=4, ensure_ascii=False)
            print(f"\nJSON output saved to: {output_path}")

        except json.JSONDecodeError:
            print("Raw output (failed to parse JSON):")
            print(content)
        
    except Exception as e:
        print(f"Error calling Qwen VL: {e}")

if __name__ == "__main__":
    summarize_scenes()
