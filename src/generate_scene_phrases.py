import os
import sys
import json
import csv
import openpyxl
import google.generativeai as genai
from openai import OpenAI
import typing_extensions as typing
import PIL.Image
import requests
import io
import datetime

# Add project root to sys.path to import src
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
sys.path.append(project_root)

from src.utils import load_config, call_qwen

def get_product_info(file_path):
    """
    Reads product name, detail, and image path from ec_product.xlsx (or csv).
    Assumes columns 'name', 'detail', and 'image' exist.
    """
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return None, None, None

    name = None
    detail = None
    image_path = None

    try:
        if file_path.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    name = row.get('name', '').strip()
                    detail = row.get('detail', '')
                    image_path = row.get('image', '').strip()
                    return name, detail, image_path
        elif file_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            
            headers = {}
            for cell in sheet[1]:
                if cell.value:
                    headers[cell.column] = cell.value
            
            name_col = None
            detail_col = None
            image_col = None
            
            for col_idx, header_val in headers.items():
                if header_val == 'name':
                    name_col = col_idx
                elif header_val == 'detail':
                    detail_col = col_idx
                elif header_val == 'image':
                    image_col = col_idx
            
            if not name_col: print("Column 'name' not found in xlsx")
            if not detail_col: print("Column 'detail' not found in xlsx")
            if not image_col: print("Column 'image' not found in xlsx")

            if sheet.max_row >= 2:
                if name_col:
                    name = sheet.cell(row=2, column=name_col).value
                if detail_col:
                    detail = sheet.cell(row=2, column=detail_col).value
                if image_col:
                    image_path = sheet.cell(row=2, column=image_col).value
                return name, detail, image_path
            else:
                print("No data rows in xlsx")
                
    except Exception as e:
        print(f"Error reading file: {e}")
        return None, None, None
    
    return None, None, None

def call_gemini(api_key, prompt, model_name="gemini-3-flash-preview"):
    """
    Calls Gemini model to generate scene phrases.
    """
    genai.configure(api_key=api_key)
    
    class Scene(typing.TypedDict):
        scene_no: int
        scene_description: str

    class Scenes(typing.TypedDict):
        scenes: list[Scene]

    model = genai.GenerativeModel(
        model_name,
        generation_config=genai.GenerationConfig(
            response_mime_type="application/json",
            response_schema=Scenes
        )
    )

    try:
        response = model.generate_content(
            contents=[
                {"role": "user", "parts": ["你是一位资深电商运营与海报摄影导演。", prompt]}
            ]
        )
        return json.loads(response.text)
    except Exception as e:
        print(f"Error calling Gemini: {e}")
        return None

def generate_images_with_gemini(api_key, image_path, scenes, output_dir, prefix="generated"):
    """
    Generates images using Gemini based on scene descriptions.
    """
    if not image_path:
        print("No image path provided.")
        return

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    genai.configure(api_key=api_key)
    model_name = "gemini-2.5-flash-image" 
    
    try:
        model = genai.GenerativeModel(model_name)
    except Exception as e:
        print(f"Error initializing model {model_name}: {e}")
        return

    try:
        if image_path.startswith('http'):
            print(f"Downloading image from {image_path}...")
            response = requests.get(image_path)
            response.raise_for_status()
            original_image = PIL.Image.open(io.BytesIO(response.content))
        else:
            if not os.path.exists(image_path):
                 print(f"Image file not found: {image_path}")
                 return
            original_image = PIL.Image.open(image_path)
    except Exception as e:
        print(f"Error opening image {image_path}: {e}")
        return
    
    print(f"\nStarting image generation using {model_name} (Prefix: {prefix})...")

    for i, scene in enumerate(scenes):
        scene_desc = scene.get("scene_description", "")
        if not scene_desc:
            continue
            
        prompt = f"把图像中的商品，放在带有{scene_desc}的场景中，商品占比不低于75%。综合以上生成一张商品展示图。"
        print(f"Generating image {i+1}/{len(scenes)}: {prompt[:50]}...")
        
        try:
            response = model.generate_content([prompt, original_image])
            
            image_saved = False
            try:
                # Check for candidates with inline data (standard for Gemini/Imagen image responses)
                if response.candidates and response.candidates[0].content.parts:
                    for part in response.candidates[0].content.parts:
                        if part.inline_data and part.inline_data.data:
                            output_filename = f"{prefix}_image_{i+1}.png"
                            output_path = os.path.join(output_dir, output_filename)
                            with open(output_path, "wb") as f:
                                f.write(part.inline_data.data)
                            print(f"  > Saved to {output_path}")
                            image_saved = True
                            break
            except Exception as e:
                 pass

            if not image_saved:
                # Fallback: Sometimes response.text contains the image description if generation failed
                # or if the model behaves differently.
                print(f"  > No image data found. Response: {response.text if hasattr(response, 'text') else 'No text'}")
                
        except Exception as e:
             print(f"  > Error generating image: {e}")

def generate_phrases():
    config = load_config()
    
    qwen_api_key = config.get("api_keys", {}).get("dashscope") or os.environ.get("DASHSCOPE_API_KEY")
    gemini_api_key = config.get("api_keys", {}).get("gemini") or os.environ.get("GEMINI_API_KEY")

    if not qwen_api_key: print("Error: DASHSCOPE_API_KEY not found.")
    if not gemini_api_key: print("Error: GEMINI_API_KEY not found.")

    file_path = os.path.join(project_root, "ec_product.xlsx")
    product_name, product_function, image_path = get_product_info(file_path)
    
    if not product_name:
        print("Error: Could not read product info from file.")
        return

    print(f"Product Name: {product_name}")
    print(f"Product Function: {product_function}")
    print(f"Product Image: {image_path}")

    image_num = 3

    prompt_template = """
### Role
你是一位资深电商运营与海报摄影导演，将帮助我生成商品展示图的“场景短语”。 

### Target
在保证原始商品照片（外观、色彩、形状）高度还原的前提下，生成一个适合电商平台的展示场景描述。 

### Sample Prompt
把图像中的商品，放在带有{{}}的场景中，商品占比不低于75%。综合以上生成一张商品展示图。

### Task
请根据以下信息输出 {image_num} 条场景短语，用于填充Sample Prompt中的 {{}}： 
- 产品名称：{product_name} 
- 产品描述：{product_function} 
 
### Workflow 
1. 先学习产品名称与产品描述，提炼 1 个最核心用途 + 2–3 个最常见真实使用地点（只选高频、可落地的场景，根据你的理解，适当泛化）。
2. 场景强约束：每条场景短语必须是该产品最常见的真实使用地点/使用环境的静态布景，禁止通用影棚背景、抽象背景、纯装饰无场景属性背景。
3. 词汇结构强约束：每条短语必须包含
    a) 1 个地点词（明确在哪里）
    b) 1 个关键环境元素词（能证明这是该场景的物件/材质/区域）
    同时轻道具总数 ≤ 2。
2. 每条场景短语需描述环境/地点/季节氛围/材质基底/光线质感/画面风格。不要使用任何操作动词或步骤描述，不出现人物/手、箭头、编号或说明文字。 
3. 场景应与产品作用形成自然联想，但不展示产品的使用过程。 
4. 背景干净简洁，商业摄影棚质感。 
5. 每条短语控制在 20–40 字之间，使用中文。
6. 补充说明：
    - detail (细节展示)：描述产品在场景中的具体视觉细节，如光影、质感、纹理等。
    - selling_point (卖点展示)：用画面语言描述产品的核心卖点或功效，避免抽象的功能性文字，转化为视觉可感知的元素。
"""

    prompt = prompt_template.format(product_name=product_name, product_function=product_function, image_num=image_num)
    
    qwen_scenes = []
    gemini_scenes = []

    if qwen_api_key:
        print("\nRunning Qwen...")
        client = OpenAI(
            api_key=qwen_api_key,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
        )
        messages = [
            {"role": "system", "content": "你是一位资深电商运营与海报摄影导演。"},
            {"role": "user", "content": prompt}
        ]
        tools = [
            {
                "type": "function",
                "function": {
                    "name": "generate_scene_phrases",
                    "description": f"生成 {image_num} 条场景短语",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "scenes": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "scene_no": {"type": "integer"},
                                        "scene_description": {"type": "string"}
                                    },
                                    "required": ["scene_no", "scene_description"]
                                }
                            }
                        },
                        "required": ["scenes"]
                    }
                }
            }
        ]

        result = call_qwen(client, "qwen-plus", messages, tools=tools, tool_name="generate_scene_phrases")
        
        if result:
            final_result = result
            if isinstance(result, dict) and "scenes" in result:
                 final_result = result["scenes"]
            
            qwen_scenes = final_result
            print("Qwen Result:", json.dumps(final_result, ensure_ascii=False, indent=2))
            
            output_path = os.path.join(os.path.dirname(__file__), "scene_phrases_qwen.json")
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(final_result, f, ensure_ascii=False, indent=2)
        else:
            print("Failed to generate scenes with Qwen.")

    if gemini_api_key:
        print("\nRunning Gemini (Text Generation)...")
        result = call_gemini(gemini_api_key, prompt)
        if result:
            final_result = result
            if isinstance(result, dict) and "scenes" in result:
                 final_result = result["scenes"]
            
            gemini_scenes = final_result
            print("Gemini Result:", json.dumps(final_result, ensure_ascii=False, indent=2))
            output_path = os.path.join(os.path.dirname(__file__), "scene_phrases_gemini.json")
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(final_result, f, ensure_ascii=False, indent=2)

    # --- Image Generation ---
    if not image_path:
        print("\nSkipping image generation (Missing image path).")
        return

    # Create a unique output directory with timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir_name = f"run_{timestamp}"
    output_dir = os.path.join(project_root, "outputs", output_dir_name)
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    print(f"\nSaving images to: {output_dir}")
    
    if qwen_scenes:
         print("\nGenerating images from Qwen scenes...")
         generate_images_with_gemini(gemini_api_key, image_path, qwen_scenes, output_dir, prefix="qwen")
    else:
         print("\nSkipping Qwen image generation (No Qwen scenes).")

    if gemini_scenes:
         print("\nGenerating images from Gemini scenes...")
         generate_images_with_gemini(gemini_api_key, image_path, gemini_scenes, output_dir, prefix="gemini")
    else:
         print("\nSkipping Gemini image generation (No Gemini scenes).")

if __name__ == "__main__":
    generate_phrases()
